from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import datetime, timedelta
from typing import Any

from django.db import transaction
from django.db.models import Count
from django.utils import timezone

from .models import AttendanceRecord, Student

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency
    load_workbook = None


@dataclass(frozen=True)
class FailedImportRow:
    row_number: int
    raw: dict[str, Any]
    errors: list[str]


@dataclass(frozen=True)
class ImportSummary:
    total_records: int
    successfully_imported: int
    duplicates_skipped: int
    invalid_rows: int
    failed_rows: list[FailedImportRow]


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_csv_rows(uploaded_file) -> list[tuple[int, dict[str, Any]]]:
    raw = uploaded_file.read()
    if isinstance(raw, bytes):
        raw = raw.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(raw))
    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, row in enumerate(reader, start=2):
        rows.append((row_number, dict(row)))
    return rows


def _parse_xlsx_rows(uploaded_file) -> list[tuple[int, dict[str, Any]]]:
    if load_workbook is None:
        raise RuntimeError("Excel import requires openpyxl. Add it to the environment to enable .xlsx support.")

    workbook = load_workbook(io.BytesIO(uploaded_file.read()), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_clean_text(header) for header in rows[0]]
    parsed_rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, values in enumerate(rows[1:], start=2):
        parsed_rows.append((row_number, {headers[index]: values[index] if index < len(values) else None for index in range(len(headers))}))
    return parsed_rows


def parse_import_rows(uploaded_file) -> list[tuple[int, dict[str, Any]]]:
    file_name = (getattr(uploaded_file, "name", "") or "").lower()
    if file_name.endswith(".xlsx"):
        return _parse_xlsx_rows(uploaded_file)
    return _parse_csv_rows(uploaded_file)


def import_students(uploaded_file) -> ImportSummary:
    rows = parse_import_rows(uploaded_file)
    total_records = len(rows)
    if not rows:
        return ImportSummary(0, 0, 0, 0, [])

    requested_indexes = [_clean_text(row.get("Index Number") or row.get("index_number")) for _, row in rows]
    existing_indexes = set(
        Student.objects.filter(index_number__in=[value for value in requested_indexes if value]).values_list("index_number", flat=True)
    )

    seen_indexes: set[str] = set()
    failed_rows: list[FailedImportRow] = []
    rows_to_create: list[dict[str, Any]] = []
    duplicates_skipped = 0
    invalid_rows = 0

    for row_number, raw_row in rows:
        full_name = _clean_text(raw_row.get("Full Name") or raw_row.get("full_name") or raw_row.get("name"))
        index_number = _clean_text(raw_row.get("Index Number") or raw_row.get("index_number") or raw_row.get("index"))
        department = _clean_text(raw_row.get("Department") or raw_row.get("department"))
        level = _clean_text(raw_row.get("Level") or raw_row.get("level"))

        errors: list[str] = []
        if not full_name:
            errors.append("Full name is required.")
        if not index_number:
            errors.append("Index number is required.")

        if index_number and index_number in seen_indexes:
            duplicates_skipped += 1
            failed_rows.append(FailedImportRow(row_number=row_number, raw=raw_row, errors=["Duplicate index number in uploaded file."]))
            continue

        if index_number and index_number in existing_indexes:
            duplicates_skipped += 1
            failed_rows.append(FailedImportRow(row_number=row_number, raw=raw_row, errors=["Student already exists in the database."]))
            continue

        if errors:
            invalid_rows += 1
            failed_rows.append(FailedImportRow(row_number=row_number, raw=raw_row, errors=errors))
            continue

        seen_indexes.add(index_number)
        rows_to_create.append(
            {
                "full_name": full_name,
                "index_number": index_number,
                "department": department or None,
                "level": level or None,
            }
        )

    successfully_imported = 0
    with transaction.atomic():
        for payload in rows_to_create:
            Student.objects.create(**payload)
            successfully_imported += 1

    return ImportSummary(
        total_records=total_records,
        successfully_imported=successfully_imported,
        duplicates_skipped=duplicates_skipped,
        invalid_rows=invalid_rows,
        failed_rows=failed_rows,
    )


def _latest_attendance_by_student():
    records = AttendanceRecord.objects.select_related("student").order_by("student_id", "-timestamp", "-id")
    latest: dict[int, AttendanceRecord] = {}
    for record in records:
        if record.student_id not in latest:
            latest[record.student_id] = record
    return latest


def get_current_students_inside() -> list[Student]:
    latest_records = _latest_attendance_by_student()
    student_ids = [record.student_id for record in latest_records.values() if record.action == AttendanceRecord.Action.ENTRY]
    if not student_ids:
        return []
    inside_students = Student.objects.filter(id__in=student_ids).order_by("full_name", "index_number")
    return list(inside_students)


def get_attendance_snapshot() -> dict[str, Any]:
    students_inside = get_current_students_inside()
    today = timezone.localdate()
    start = timezone.make_aware(datetime.combine(today, datetime.min.time()))
    end = start + timedelta(days=1)
    qs = AttendanceRecord.objects.filter(timestamp__gte=start, timestamp__lt=end)
    return {
        "occupancy": len(students_inside),
        "students_inside": students_inside,
        "entries_today": qs.filter(action=AttendanceRecord.Action.ENTRY).count(),
        "exits_today": qs.filter(action=AttendanceRecord.Action.EXIT).count(),
    }


def get_student_attendance_history(student: Student | None = None, limit: int = 25):
    qs = AttendanceRecord.objects.select_related("student")
    if student is not None:
        qs = qs.filter(student=student)
    return qs.order_by("-timestamp", "-id")[:limit]


def record_student_action(student: Student, action: str) -> tuple[AttendanceRecord | None, str]:
    last_record = AttendanceRecord.objects.filter(student=student).order_by("-timestamp", "-id").first()

    if action == AttendanceRecord.Action.ENTRY and last_record and last_record.action == AttendanceRecord.Action.ENTRY:
        return None, "This student already has an active entry."

    if action == AttendanceRecord.Action.EXIT and (not last_record or last_record.action == AttendanceRecord.Action.EXIT):
        return None, "This student cannot exit before a matching entry."

    record = AttendanceRecord.objects.create(student=student, action=action)
    return record, "Attendance recorded successfully."


def get_student_analytics(limit: int = 5) -> dict[str, Any]:
    departments = (
        Student.objects.exclude(department__isnull=True)
        .exclude(department__exact="")
        .values("department")
        .annotate(total=Count("id"))
        .order_by("-total", "department")
    )
    top_students = (
        AttendanceRecord.objects.select_related("student")
        .values("student_id", "student__full_name", "student__index_number")
        .annotate(total=Count("id"))
        .order_by("-total", "student__full_name")[:limit]
    )
    return {
        "departments": list(departments),
        "top_students": list(top_students),
    }
